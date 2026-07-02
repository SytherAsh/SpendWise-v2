"""
Step 2 of the labeling pipeline. Reads the hand-filled review_sheet.xlsx,
joins those labels back onto the unmatched rows, combines with
auto_labeled.csv into datasets/<name>/labeled.csv, and — this is the point of
the whole pipeline — appends the newly-labeled recipient patterns into
knowledge_base/merchant_rules.csv so every future dataset needs less manual
review than this one did.

Usage:
    python apply_review.py --dataset spendwise_4yr_sbi_statement
"""

import argparse
import csv
import sys
from datetime import date
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from categories import CANONICAL_FIELDS
from dataset_config import get_dataset

LABELING_ROOT = Path(__file__).parent.parent
KB_DIR = LABELING_ROOT / "knowledge_base"
MERCHANT_RULES_PATH = KB_DIR / "merchant_rules.csv"


def load_existing_patterns(path):
    patterns = set()
    if not path.exists():
        return patterns
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            patterns.add((row["pattern"].lower(), row["match_field"]))
    return patterns


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", required=True)
    args = parser.parse_args()

    config = get_dataset(args.dataset)
    dataset_dir = LABELING_ROOT / "datasets" / args.dataset
    review_path = dataset_dir / "review_sheet.xlsx"
    auto_labeled_path = dataset_dir / "auto_labeled.csv"

    if not review_path.exists():
        print(f"No review_sheet.xlsx found for '{args.dataset}' — run auto_label.py first.")
        sys.exit(1)

    # --- re-derive the unmatched rows from the raw file, keyed by recipient ---
    raw_path = dataset_dir / config["raw_file"]
    wb = openpyxl.load_workbook(raw_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers, data_rows = rows[0], rows[1:]

    already_labeled_recipients = set()
    auto_rows = []
    if auto_labeled_path.exists():
        with open(auto_labeled_path, newline="", encoding="utf-8") as f:
            auto_rows = list(csv.DictReader(f))
            already_labeled_recipients = {r["recipient_name"] for r in auto_rows}

    review_wb = openpyxl.load_workbook(review_path, read_only=True)
    review_ws = review_wb.active
    recipient_to_category = {}
    unlabeled_recipients = []
    for name, count, upi, note, mode, category in review_ws.iter_rows(min_row=2, values_only=True):
        if not category:
            unlabeled_recipients.append(name)
            continue
        recipient_to_category[name] = category

    if unlabeled_recipients:
        print(f"WARNING: {len(unlabeled_recipients)} recipient group(s) left blank — "
              f"their rows will be skipped: {unlabeled_recipients[:10]}"
              f"{'...' if len(unlabeled_recipients) > 10 else ''}")

    reviewed_rows = []
    for raw_row in data_rows:
        row_dict = dict(zip(headers, raw_row))
        recipient = row_dict.get("Recipient_Name")
        if recipient in already_labeled_recipients or recipient not in recipient_to_category:
            continue
        canonical = {field: None for field in CANONICAL_FIELDS}
        for raw_col, value in row_dict.items():
            target = config["column_map"].get(raw_col, "__unmapped__")
            if target:
                canonical[target] = value
        canonical["source"] = config["source_label"]
        canonical["category"] = recipient_to_category[recipient]
        reviewed_rows.append(canonical)

    # --- write combined labeled.csv ---
    out_fields = CANONICAL_FIELDS + ["category"]
    with open(dataset_dir / "labeled.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=out_fields)
        writer.writeheader()
        writer.writerows(auto_rows)
        writer.writerows(reviewed_rows)

    # --- grow the knowledge base with the newly reviewed recipients ---
    existing_patterns = load_existing_patterns(MERCHANT_RULES_PATH)
    today = date.today().isoformat()
    new_rules = []
    for name, category in recipient_to_category.items():
        key = (str(name).lower(), "recipient_name")
        if key in existing_patterns:
            continue
        existing_patterns.add(key)
        new_rules.append({
            "pattern": name,
            "match_field": "recipient_name",
            "category": category,
            "source_dataset": args.dataset,
            "added_date": today,
        })

    if new_rules:
        write_header = not MERCHANT_RULES_PATH.exists() or MERCHANT_RULES_PATH.stat().st_size == 0
        with open(MERCHANT_RULES_PATH, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["pattern", "match_field", "category", "source_dataset", "added_date"])
            if write_header:
                writer.writeheader()
            writer.writerows(new_rules)

    print(f"{args.dataset}: wrote {len(auto_rows) + len(reviewed_rows)} labeled rows to labeled.csv "
          f"({len(auto_rows)} auto + {len(reviewed_rows)} reviewed)")
    print(f"  knowledge base grew by {len(new_rules)} merchant rule(s)")


if __name__ == "__main__":
    main()
