"""
Step 1 of the labeling pipeline. Normalizes a raw dataset to the canonical
schema, applies the knowledge_base rules, and splits the result into rows
that were auto-labeled and rows that still need a human to look at them.

Usage:
    python auto_label.py --dataset spendwise_4yr_sbi_statement

Outputs (under datasets/<name>/):
    auto_labeled.csv   rows the knowledge base already resolved
    review_sheet.xlsx  unresolved rows, grouped by recipient_name, with a
                        Category dropdown — hand this to a human

Safe to re-run: as the knowledge base grows, coverage on the same raw file
only goes up.
"""

import argparse
import csv
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
from categories import CATEGORIES, CANONICAL_FIELDS
from dataset_config import get_dataset

LABELING_ROOT = Path(__file__).parent.parent
KB_DIR = LABELING_ROOT / "knowledge_base"


def load_rules(path, fields):
    rules = []
    if not path.exists():
        return rules
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get(fields[0]):
                rules.append(row)
    return rules


def normalize_row(raw_row, headers, column_map, source_label):
    canonical = {field: None for field in CANONICAL_FIELDS}
    for raw_col, value in zip(headers, raw_row):
        target = column_map.get(raw_col, "__unmapped__")
        if target:
            canonical[target] = value
    canonical["source"] = source_label
    return canonical


def match_merchant(row, merchant_rules):
    recipient = str(row.get("recipient_name") or "").lower()
    upi = str(row.get("upi_id") or "").lower()
    for rule in merchant_rules:
        pattern = rule["pattern"].lower()
        field = rule["match_field"]
        if field in ("recipient_name", "both") and pattern in recipient:
            return rule["category"]
        if field in ("upi_id", "both") and pattern in upi:
            return rule["category"]
    return None


def match_keyword(row, keyword_rules):
    note = str(row.get("note") or "").lower()
    if not note:
        return None
    for rule in keyword_rules:
        if rule["pattern"].lower() in note:
            return rule["category"]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", required=True)
    args = parser.parse_args()

    config = get_dataset(args.dataset)
    if config["source_type"] != "bank_statement":
        print(f"Dataset '{args.dataset}' has source_type={config['source_type']!r}, "
              f"which is not directly labelable yet (needs a parsing step first). See "
              f"dataset_config.py and ml/labeling/tracking/LABELING_STATUS.md.")
        sys.exit(1)

    dataset_dir = LABELING_ROOT / "datasets" / args.dataset
    raw_path = dataset_dir / config["raw_file"]

    wb = openpyxl.load_workbook(raw_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers, data_rows = rows[0], rows[1:]

    merchant_rules = load_rules(KB_DIR / "merchant_rules.csv", ["pattern"])
    keyword_rules = load_rules(KB_DIR / "keyword_rules.csv", ["pattern"])

    auto_labeled = []
    unmatched_groups = defaultdict(list)

    for raw_row in data_rows:
        canonical = normalize_row(raw_row, headers, config["column_map"], config["source_label"])
        category = match_merchant(canonical, merchant_rules) or match_keyword(canonical, keyword_rules)
        if category:
            canonical["category"] = category
            auto_labeled.append(canonical)
        else:
            unmatched_groups[canonical.get("recipient_name")].append(canonical)

    # --- write auto_labeled.csv ---
    out_fields = CANONICAL_FIELDS + ["category"]
    with open(dataset_dir / "auto_labeled.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=out_fields)
        writer.writeheader()
        writer.writerows(auto_labeled)

    # --- write review_sheet.xlsx, grouped by recipient, sorted by frequency ---
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Review"
    out_ws.append(["Recipient_Name", "Count", "Sample_UPI_ID", "Sample_Note", "Sample_Mode", "Category"])

    for name, items in sorted(unmatched_groups.items(), key=lambda kv: -len(kv[1])):
        sample = items[0]
        out_ws.append([
            name,
            len(items),
            sample.get("upi_id"),
            sample.get("note"),
            sample.get("transaction_mode"),
            "",
        ])

    dv = DataValidation(type="list", formula1='"' + ",".join(CATEGORIES) + '"', allow_blank=True)
    out_ws.add_data_validation(dv)
    dv.add(f"F2:F{out_ws.max_row}")
    for col, width in zip("ABCDEF", [20, 8, 16, 24, 14, 20]):
        out_ws.column_dimensions[col].width = width

    out_wb.save(dataset_dir / "review_sheet.xlsx")

    total = len(data_rows)
    coverage = len(auto_labeled) / total * 100 if total else 0
    print(f"{args.dataset}: {total} rows total")
    print(f"  auto-labeled: {len(auto_labeled)} ({coverage:.1f}%)")
    print(f"  needs review: {sum(len(v) for v in unmatched_groups.values())} rows "
          f"across {len(unmatched_groups)} recipient groups -> review_sheet.xlsx")


if __name__ == "__main__":
    main()
